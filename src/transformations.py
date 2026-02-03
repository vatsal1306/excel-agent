import datetime
import datetime as dt
import math
import os.path
import re
from copy import copy
from io import BytesIO
from typing import Optional, Tuple, Union, List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas.api.types import is_datetime64_any_dtype
from tqdm.auto import tqdm

from src import OUTPUT_ROOT
from src.Logging import logger


def step_01(df: pd.DataFrame, save: bool = False) -> pd.DataFrame:
    """ STEP 1 """

    cols_to_delete = [
        'Sales Office', 'Sales Group', 'Document Date', 'Payer', 'Name of additional partner',
        'Incoterms', 'Plant Name', 'Open Indicator', 'MSP Category', 'Material', 'Usage Description',
        'Gross weight', 'Weight unit'
    ]

    for col in cols_to_delete:
        try:
            df.drop(columns=[col], inplace=True)
        except KeyError:
            logger.error(f"Column not found, skipping: {col}")

    # --- Force "Last G/I Date" to TEXT DD/MM/YYYY (do NOT keep datetime dtype) ---
    date_col = "Last G/I Date"
    if date_col in df.columns:
        s = df[date_col]

        # Case 1: already a pandas datetime dtype
        if is_datetime64_any_dtype(s):
            df[date_col] = s.dt.strftime("%m/%d/%Y")
        else:
            # Case 2: mixed/object column – normalize any datetime-like objects
            def _fmt_one(x):
                if x is None or (isinstance(x, float) and pd.isna(x)):
                    return ""
                if isinstance(x, pd.Timestamp):
                    x = x.to_pydatetime()
                if isinstance(x, datetime.datetime):
                    d = x.date()
                    return f"{d.month:02d}/{d.day:02d}/{d.year}"
                if isinstance(x, datetime.date):
                    return f"{x.month:02d}/{x.day:02d}/{x.year}"

                # Case 3: strings like "2025-03-26 00:00:00" or other parseable variants
                if isinstance(x, str):
                    t = x.strip()
                    if t == "":
                        return ""
                    parsed = pd.to_datetime(t, dayfirst=True, errors="coerce")
                    if pd.notna(parsed):
                        d = parsed.date()
                        return f"{d.month:02d}/{d.day:02d}/{d.year}"
                    return t  # leave non-date strings as-is

                # fallback: stringify
                return str(x)

            df[date_col] = s.apply(_fmt_one).astype(str)

        # Ensure blanks don't show as "nan"/"NaT"
        df[date_col] = df[date_col].replace({"nan": "", "NaT": ""})

    else:
        logger.warning("Column 'Last G/I Date' not found; no date coercion applied.")

    logger.info(f"Remaining columns after deletion: {list(df.columns)}")

    if save:
        df.to_excel(os.path.join(OUTPUT_ROOT, 'output_step1.xlsx'), index=False)

    return df


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().casefold()


def _to_float(v):
    """Convert Excel cell values to float when possible; return None for blanks/non-numeric."""
    if v is None:
        return None
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        try:
            if v != v:  # NaN
                return None
        except Exception:
            pass
        return float(v)
    if isinstance(v, str):
        t = v.strip()
        if t == "":
            return None
        t = re.sub(r"[,\u20B9$]", "", t)  # remove commas/currency symbols
        try:
            return float(t)
        except Exception:
            return None
    return None


def _excel_date_to_mmddyyyy_string(v) -> str:
    """
    Convert various Excel/openpyxl date representations to 'MM/DD/YYYY' string.
    Keeps non-date strings as-is (stringified).
    """
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        d = v.date()
        return f"{d.month:02d}/{d.day:02d}/{d.year}"
    if isinstance(v, datetime.date):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if isinstance(v, float) and math.isnan(v):
                return ""
        except Exception:
            pass
        try:
            d = from_excel(v)
            if isinstance(d, datetime.datetime):
                d = d.date()
            if isinstance(d, datetime.date):
                return f"{d.month:02d}/{d.day:02d}/{d.year}"
        except Exception:
            pass
    return str(v)


def step_02(
        file_in: Union[str, os.PathLike, pd.DataFrame],
        sheet_name: str | None = None,
        header_scan_rows: int = 20,
        keep_net_value_blanks: bool = True,
        save: bool = False,
):
    """STEP 2 (new order):
    1) Invoice Quantity – Only keep the zeros checked, then hide
    2) Net Value – uncheck all zeros and hide
    3) Hide MSP Invoice Quantity
    """

    # ---- Load workbook ----
    if isinstance(file_in, pd.DataFrame):
        df = file_in.copy()

        # Keep 'Last G/I Date' as text DD/MM/YYYY before writing to Excel
        if "Last G/I Date" in df.columns:
            s = df["Last G/I Date"]
            if pd.api.types.is_datetime64_any_dtype(s):
                df["Last G/I Date"] = s.dt.strftime("%d/%m/%Y")
            else:
                df["Last G/I Date"] = (
                    s.astype(str)
                    .replace("NaT", "")
                    .replace("nan", "")
                )

        excel_stream = BytesIO()
        df.to_excel(
            excel_stream,
            index=False,
            sheet_name=sheet_name or "Sheet1",
            engine="openpyxl",
        )
        excel_stream.seek(0)
        wb = load_workbook(excel_stream)
    elif isinstance(file_in, (str, os.PathLike)):
        wb = load_workbook(file_in)
    else:
        raise TypeError("file_in must be a path to .xlsx (str/PathLike) or a pandas DataFrame")

    ws = wb[sheet_name] if sheet_name else wb.active

    # Columns we must find (also used for hiding)
    targets = ["Net Value of Confirmed Quantity", "Invoice Qty", "MSP Invoice Qty"]

    # ---- Find header row (best match among first N rows) ----
    best_row, best_hits, best_map = None, -1, {}
    for r in range(1, min(header_scan_rows, ws.max_row) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        col_map = {}
        for c, v in enumerate(row_vals, start=1):
            nv = _norm(v)
            for t in targets:
                if nv == _norm(t):
                    col_map[t] = c
        if len(col_map) > best_hits:
            best_hits, best_row, best_map = len(col_map), r, col_map

    if not best_row or best_hits == 0:
        raise ValueError("Could not locate header row / required columns in the first scanned rows.")

    header_row = best_row

    # ---- Force "Last G/I Date" to remain TEXT in DD/MM/YYYY ----
    date_col = None
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(header_row, c).value) == _norm("Last G/I Date"):
            date_col = c
            break

    if date_col is None:
        logger.warning("Missing column: 'Last G/I Date' (no date coercion applied).")
    else:
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(r, date_col)
            cell.value = _excel_date_to_mmddyyyy_string(cell.value)
            cell.number_format = "@"  # force text display :contentReference[oaicite:2]{index=2}

    # ---- Reset hidden rows so script is idempotent ----
    for r in range(1, ws.max_row + 1):
        if ws.row_dimensions[r].hidden:
            ws.row_dimensions[
                r].hidden = False  # hide/unhide via row_dimensions.hidden :contentReference[oaicite:3]{index=3}

    net_col = best_map.get("Net Value of Confirmed Quantity")
    inv_col = best_map.get("Invoice Qty")

    # ---- Apply "filters" by hiding rows (deterministic) ----
    # NEW ORDER:
    # 1) Invoice Qty -> keep only zeros (hide everything else)
    # 2) Net Value -> exclude zeros (hide zeros) on remaining visible rows
    inv_hidden = 0
    net_hidden = 0

    # 1) Invoice Qty FIRST
    if inv_col is None:
        logger.error("Missing column: 'Invoice Qty' (skipping filter 1).")
    else:
        for r in range(header_row + 1, ws.max_row + 1):
            v = _to_float(ws.cell(r, inv_col).value)
            if v != 0.0:  # includes None and non-zero
                ws.row_dimensions[r].hidden = True
                inv_hidden += 1

    # 2) Net Value SECOND
    if net_col is None:
        logger.error("Missing column: 'Net Value of Confirmed Quantity' (skipping filter 2).")
    else:
        for r in range(header_row + 1, ws.max_row + 1):
            if ws.row_dimensions[r].hidden:
                continue  # already excluded by Invoice filter

            v = _to_float(ws.cell(r, net_col).value)
            if v == 0.0:
                ws.row_dimensions[r].hidden = True
                net_hidden += 1
            elif v is None and not keep_net_value_blanks:
                ws.row_dimensions[r].hidden = True
                net_hidden += 1

    logger.info(f"Rows hidden by Invoice Qty filter: {inv_hidden}")
    logger.info(f"Rows hidden by Net Value filter: {net_hidden}")
    logger.info(f"Total rows hidden by criteria: {inv_hidden + net_hidden}")

    # ---- Ensure filter dropdowns exist in Excel UI ----
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"  # :contentReference[oaicite:4]{index=4}

    # ---- Hide specified columns ----
    for col_name in targets:
        c = best_map.get(col_name)
        if c is None:
            logger.warning(f"Missing column: '{col_name}' (cannot hide).")
            continue
        ws.column_dimensions[get_column_letter(c)].hidden = True  # :contentReference[oaicite:5]{index=5}

    if save:
        wb.save(os.path.join(OUTPUT_ROOT, "output_step2.xlsx"))

    return wb


def _parse_date(v) -> Optional[dt.date]:
    """Parse many Excel-ish date representations into dt.date (for sorting only).
    Accepts DD/MM/YYYY and MM/DD/YYYY; if ambiguous, assumes DD/MM/YYYY.
    """
    if v is None or v == "":
        return None

    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v

    # Excel serial numbers sometimes come in as int/float
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if isinstance(v, float) and math.isnan(v):
                return None
        except Exception:
            pass
        try:
            d = from_excel(v)
            return d.date() if isinstance(d, dt.datetime) else d
        except Exception:
            return None

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None

        # Strip common time suffix if present (prevents timestamp issues affecting parsing)
        # e.g. "2026-02-04 00:00:00" -> "2026-02-04"
        s = re.split(r"\s+", s, maxsplit=1)[0]

        # Handle slash dates with ambiguity rules
        m = re.match(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$", s)
        if m:
            a = int(m.group(1))
            b = int(m.group(2))
            y = int(m.group(3))

            # If a > 12 -> must be DD/MM/YYYY
            if a > 12:
                d, mo = a, b
            # If b > 12 -> must be MM/DD/YYYY
            elif b > 12:
                mo, d = a, b
            # Ambiguous -> default DD/MM/YYYY (per your requirement)
            else:
                d, mo = a, b

            try:
                return dt.date(y, mo, d)
            except Exception:
                return None

        # Non-slash known formats
        for fmt in ("%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except Exception:
                pass

        # Fallback: try dateutil if available (dayfirst=True to match your default)
        try:
            from dateutil import parser
            return parser.parse(s, dayfirst=True, fuzzy=True).date()
        except Exception:
            return None

    return None


def _date_to_mmddyyyy(d: dt.date) -> str:
    return f"{d.month:02d}/{d.day:02d}/{d.year}"


def _find_header_and_col(ws, header_text: str, header_scan_rows: int = 20) -> Tuple[int, int]:
    target = _norm(header_text)
    for r in range(1, min(header_scan_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(r, c).value) == target:
                return r, c
    raise ValueError(f"Could not find column header '{header_text}' in first {header_scan_rows} rows.")


def _snapshot_rows(ws, start_row: int, end_row: int):
    """Snapshot rows including values + style so formatting survives reordering."""
    max_col = ws.max_column
    snap = []
    for r in range(start_row, end_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            row_cells.append({
                "value": cell.value,
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
                "comment": cell.comment,
            })
        rd = ws.row_dimensions[r]
        snap.append({
            "cells": row_cells,
            "row_hidden": bool(rd.hidden),
            "row_height": rd.height,
            "outline": rd.outlineLevel,
        })
    return snap


def _write_snapshot(
        ws,
        start_row: int,
        snap,
        text_col: Optional[int] = None,
        text_col_is_date: bool = False,
        *,
        date_dayfirst: bool = True,
        date_year_len: int = 4,
):
    """Write rows back in new order; optionally force one column to TEXT strings."""
    max_col = ws.max_column
    for i, row in enumerate(snap):
        tr = start_row + i

        rd = ws.row_dimensions[tr]
        rd.hidden = row["row_hidden"]
        rd.height = row["row_height"]
        rd.outlineLevel = row["outline"]

        for c in range(1, max_col + 1):
            cell = ws.cell(tr, c)
            src = row["cells"][c - 1]
            cell._style = copy(src["style"])
            cell.number_format = src["number_format"]
            cell.value = src["value"]
            cell._hyperlink = copy(src["hyperlink"]) if src["hyperlink"] else None
            cell.comment = src["comment"]

        # Force a specific column to remain TEXT in the saved workbook
        if text_col is not None:
            dc = ws.cell(tr, text_col)
            if text_col_is_date:
                d = _parse_slash_date_with_format(dc.value, dayfirst=date_dayfirst)
                if d is None:
                    dc.value = "" if dc.value is None else str(dc.value)
                else:
                    # Always output as MM/DD/(YY|YYYY) text (no timestamp)
                    dc.value = _format_mmdd(d, year_len=date_year_len)
            else:
                dc.value = "" if dc.value is None else str(dc.value)

            dc.number_format = "@"  # Excel TEXT format


def _detect_gi_slash_format(values: list) -> tuple[bool, int]:
    """
    Detect overall slash-date format for the whole column in this run.

    Returns:
        (dayfirst, year_len)
        - dayfirst=True  => DD/MM/(YY|YYYY)
        - dayfirst=False => MM/DD/(YY|YYYY)
        - year_len is 2 or 4 (based on matched samples; defaults to 2 if only YY seen, else 4)
    Detection rule (as you asked):
        - If we see any sample where first token > 12 => DD/MM
        - If we see any sample where second token > 12 => MM/DD
        - If ambiguous across all samples => default DD/MM
    """
    ddmm_votes = 0
    mmdd_votes = 0
    year2 = 0
    year4 = 0

    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue

        # Strip time part if present
        s = re.split(r"\s+", s, maxsplit=1)[0]

        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})$", s)
        if not m:
            continue

        a = int(m.group(1))
        b = int(m.group(2))
        y = m.group(3)

        if len(y) == 2:
            year2 += 1
        else:
            year4 += 1

        if a > 12:
            ddmm_votes += 1
        elif b > 12:
            mmdd_votes += 1
        # else ambiguous -> no vote

    # Decide format
    if mmdd_votes > ddmm_votes:
        dayfirst = False
    elif ddmm_votes > mmdd_votes:
        dayfirst = True
    else:
        dayfirst = True  # ambiguous default => DD/MM

    # Decide year length (if any YYYY exists, keep 4; else 2)
    year_len = 4 if year4 > 0 else 2

    return dayfirst, year_len


def _parse_slash_date_with_format(v, *, dayfirst: bool) -> Optional[dt.date]:
    """
    Parse a date value into dt.date using a *known* dayfirst decision for this run.
    Supports DD/MM/YY, MM/DD/YY, DD/MM/YYYY, MM/DD/YYYY.
    """
    if v is None or v == "":
        return None

    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v

    # Excel serial numbers
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if isinstance(v, float) and math.isnan(v):
                return None
        except Exception:
            pass
        try:
            d = from_excel(v)
            return d.date() if isinstance(d, dt.datetime) else d
        except Exception:
            return None

    s = str(v).strip()
    if not s:
        return None

    # Strip time part if present
    s = re.split(r"\s+", s, maxsplit=1)[0]

    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})$", s)
    if not m:
        # fallback: try a couple safe formats (optional)
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", "%d.%m.%Y"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    a = int(m.group(1))
    b = int(m.group(2))
    y_raw = m.group(3)

    # interpret order based on detected format for this run
    if dayfirst:
        d, mo = a, b
    else:
        mo, d = a, b

    # Year: support YY and YYYY
    if len(y_raw) == 2:
        yy = int(y_raw)
        # Pivot rule: 00-79 => 2000-2079; 80-99 => 1980-1999
        y = 2000 + yy if yy <= 79 else 1900 + yy
    else:
        y = int(y_raw)

    try:
        return dt.date(y, mo, d)
    except Exception:
        return None


def _format_mmdd(d: dt.date, *, year_len: int) -> str:
    """Format as MM/DD/YY or MM/DD/YYYY."""
    if year_len == 2:
        return f"{d.month:02d}/{d.day:02d}/{d.year % 100:02d}"
    return f"{d.month:02d}/{d.day:02d}/{d.year}"


def step_03(input_wb: Workbook, header: str, treat_as_date: bool = False, header_scan_rows: int = 20,
            blanks_last: bool = True, save_name: str = None) -> Workbook:
    """
    Sort by ONE column, save output, and RETURN the sorted workbook so it can be chained.
    """
    bio = BytesIO()
    input_wb.save(bio)  # save workbook to memory
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active

    header_row, col = _find_header_and_col(ws, header, header_scan_rows=header_scan_rows)
    start = header_row + 1
    end = ws.max_row

    rows = _snapshot_rows(ws, start, end)

    # --- NEW: detect the input format for this whole run (only when sorting dates) ---
    date_dayfirst = True
    date_year_len = 4
    if treat_as_date:
        col_values = [r["cells"][col - 1]["value"] for r in rows]
        date_dayfirst, date_year_len = _detect_gi_slash_format(col_values)
        logger.info(
            f"Step 3 date scan: detected {'DD/MM' if date_dayfirst else 'MM/DD'} "
            f"with year_len={date_year_len}"
        )

    def key_func(row):
        v = row["cells"][col - 1]["value"]

        if treat_as_date:
            d = _parse_slash_date_with_format(v, dayfirst=date_dayfirst)
            if d is None:
                return dt.date.max if blanks_last else dt.date.min
            return d

        s = "" if v is None else str(v).strip()
        if blanks_last and s == "":
            return ("\uffff",)
        return (s.casefold(), s)

    rows_sorted = sorted(rows, key=key_func)

    _write_snapshot(
        ws,
        start,
        rows_sorted,
        text_col=col,
        text_col_is_date=treat_as_date,
        date_dayfirst=date_dayfirst,
        date_year_len=date_year_len,
    )

    if save_name:
        wb.save(os.path.join(OUTPUT_ROOT, save_name))

    return wb


def _find_header_row_best_match(ws, required_headers: List[str], header_scan_rows: int = 20) -> \
        Tuple[int, Dict[str, int]]:
    """
    Find the header row by best-match count within the first N rows.
    Returns (header_row_index, {header_name: col_index}).
    """
    required_norm = {_norm(h): h for h in required_headers}

    best_row: Optional[int] = None
    best_hits = -1
    best_map: Dict[str, int] = {}

    for r in tqdm(range(1, min(header_scan_rows, ws.max_row) + 1), desc="Finding header row"):
        hits: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            nv = _norm(v)
            if nv in required_norm:
                hits[required_norm[nv]] = c

        if len(hits) > best_hits:
            best_hits = len(hits)
            best_row = r
            best_map = hits

    if best_row is None or best_hits <= 0:
        raise ValueError("Could not locate header row in scanned range.")

    missing = [h for h in required_headers if h not in best_map]
    if missing:
        raise ValueError(f"Missing required header(s) in detected header row: {missing}")

    return best_row, best_map


def _rebuild_visible_only_sheet(
        wb: Workbook,
        ws: Worksheet,
        *,
        header_row: int,
        keep_title: bool = True,
) -> tuple[Worksheet, int, int]:
    """
    Rebuild worksheet containing ONLY visible rows/columns.
    Much faster than deleting hidden rows/cols one-by-one.

    Returns: (new_ws, rows_removed, cols_removed)
    """

    # Visible columns (by current ws state)
    visible_cols = []
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        if not bool(ws.column_dimensions[letter].hidden):
            visible_cols.append(c)

    # Visible rows: keep header row always; keep only non-hidden data rows
    visible_rows = [header_row]
    for r in range(header_row + 1, ws.max_row + 1):
        if not bool(ws.row_dimensions[r].hidden):
            visible_rows.append(r)

    rows_removed = (ws.max_row - header_row) - (len(visible_rows) - 1)
    cols_removed = ws.max_column - len(visible_cols)

    # Create temp sheet
    tmp_title = f"{ws.title}__VISIBLE__"
    if tmp_title in wb.sheetnames:
        del wb[tmp_title]
    new_ws = wb.create_sheet(tmp_title)

    # Copy column widths
    for new_c, old_c in tqdm(enumerate(visible_cols, start=1), desc="Copying column widths"):
        old_letter = get_column_letter(old_c)
        new_letter = get_column_letter(new_c)
        new_ws.column_dimensions[new_letter].width = ws.column_dimensions[old_letter].width

    # Copy rows (values + styles) and row heights
    for new_r, old_r in tqdm(enumerate(visible_rows, start=1), desc="Copying rows"):
        new_ws.row_dimensions[new_r].height = ws.row_dimensions[old_r].height
        for new_c, old_c in enumerate(visible_cols, start=1):
            _copy_cell(ws.cell(old_r, old_c), new_ws.cell(new_r, new_c))

    # Clear filters; you can re-add later if desired
    new_ws.auto_filter.ref = None

    # Replace original sheet (keep same title & position)
    old_index = wb.sheetnames.index(ws.title)
    old_title = ws.title

    wb.remove(ws)

    if keep_title:
        new_ws.title = old_title

    # Move sheet to original position
    wb._sheets.remove(new_ws)
    wb._sheets.insert(old_index, new_ws)

    return new_ws, rows_removed, cols_removed


def _safe_sheet_name(wb: Workbook, desired: str) -> str:
    """
    Excel sheet names max 31 chars and must be unique.
    If already exists, append ' (2)', ' (3)', ...
    """
    base = (desired or "Sheet").strip()
    base = base[:31] if len(base) > 31 else base

    if base not in wb.sheetnames:
        return base

    i = 2
    while True:
        suffix = f" ({i})"
        max_base = 31 - len(suffix)
        cand = (base[:max_base] + suffix) if len(base) > max_base else (base + suffix)
        if cand not in wb.sheetnames:
            return cand
        i += 1


def _copy_cell(src, dst) -> None:
    """Copy value + style-ish properties from one openpyxl cell to another."""
    dst.value = src.value
    dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst._hyperlink = copy(src.hyperlink) if src.hyperlink else None
    dst.comment = src.comment


def step_04_create_distribution_tabs(
        wb: Workbook,
        *,
        source_sheet_name: Optional[str] = None,
        header_scan_rows: int = 20,
        save: bool = False,
        save_name: str = "step4_distribution_tabs.xlsx",
) -> Workbook:
    """
    STEP 4
    - Use ONLY visible rows from Step 2 (then permanently delete hidden rows/cols).
    - Create distribution tabs by 'Name of sold-to party' (sold-to).
    - Within each tab, group by 'Name of ship-to party', sorted A→Z.
      For each contractor group:
        - Add two blank rows ABOVE (even before first group)
        - On the lower blank row, merge+center and yellow highlight across
          columns ('Name of ship-to party' .. 'Name 2'), and write contractor name.
    - If a mapped tab has no rows, do not create the sheet.
    """

    ws = wb[source_sheet_name] if source_sheet_name else wb.active

    required = ["Name of sold-to party", "Name of ship-to party", "Name 2"]
    header_row, col_map = _find_header_row_best_match(ws, required, header_scan_rows=header_scan_rows)

    ws, rows_deleted, cols_deleted = _rebuild_visible_only_sheet(wb, ws, header_row=header_row)
    logger.info(f"Step 4: rebuilt visible-only sheet. rows_removed={rows_deleted}, cols_removed={cols_deleted}")

    # Re-locate columns after deletions (indices may shift)
    header_row, col_map = _find_header_row_best_match(ws, required, header_scan_rows=header_scan_rows)
    logger.info(f"Re-detected header_row={header_row}; col_map={col_map} after deletions.")

    sold_to_col = col_map["Name of sold-to party"]
    ship_name_col = col_map["Name of ship-to party"]
    name2_col = col_map["Name 2"]

    if ws.max_row <= header_row:
        logger.warning(
            "⚠️no data rows exist below header after deleting hidden rows. "
            "No distribution sheets will be created."
        )
        if save:
            wb.save(os.path.join(OUTPUT_ROOT, save_name))
        return wb

    # Build headers list in current sheet order (post-deletion)
    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        headers.append("" if ws.cell(header_row, c).value is None else str(ws.cell(header_row, c).value))

    # Mapping: source sold-to value -> destination sheet name (your naming)
    distributor_map = {
        "ABC SUPPLY #402": "ABC East 402 (Warren)",
        "ABC SUPPLY CO INC #015": "ABC West 015 (Grand Rapids)",
        "BEACON BUILDING PRODUCTS 285": "QXO East 285 (Rochester Hills)",
        "BEACON BUILDING PRODUCTS 228": "QXO West 228 (Grand Rapids)",
        "BLOOM ROOFING SYSTEMS  INC.": "Bloom – Direct",
        "THE FRED CHRISTEN & SONS CO": "Christen Detroit – Direct",
    }
    distributor_map_norm = {_norm(k): v for k, v in distributor_map.items()}

    # Snapshot data rows (post-deletion = all rows are visible)
    data_rows: List[int] = list(range(header_row + 1, ws.max_row + 1))

    # Quick data-quality checks (log anything even slightly unexpected)
    blank_sold_to = 0
    blank_ship_to = 0
    for r in data_rows:
        if _norm(ws.cell(r, sold_to_col).value) == "":
            blank_sold_to += 1
        if _norm(ws.cell(r, ship_name_col).value) == "":
            blank_ship_to += 1

    if blank_sold_to:
        logger.warning(f"⚠️found {blank_sold_to} row(s) with blank 'Name of sold-to party'.")
    if blank_ship_to:
        logger.warning(f"⚠️found {blank_ship_to} row(s) with blank 'Name of ship-to party' (contractor name).")

    # Pre-snapshot row content for speed + stable output
    row_snaps: Dict[int, List] = {}
    for i, r in enumerate(data_rows, start=1):
        row_snaps[r] = [ws.cell(r, c) for c in range(1, ws.max_column + 1)]
        if i % 5000 == 0 or i == len(data_rows):
            logger.info(f"snapshotted {i}/{len(data_rows)} rows.")

    # Copy column widths from source sheet
    src_col_widths: Dict[int, Optional[float]] = {}
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        src_col_widths[c] = ws.column_dimensions[letter].width

    # (Optional) warn if ship-to and name2 columns are not adjacent (not wrong, but good to know)
    if abs(ship_name_col - name2_col) != 1:
        logger.warning(
            f"⚠️'Name of ship-to party' (col={ship_name_col}) and 'Name 2' (col={name2_col}) "
            f"are not adjacent; merge will span columns {min(ship_name_col, name2_col)}..{max(ship_name_col, name2_col)}."
        )

    # Log sold-to distribution observed vs expected mapping
    observed_sold_to = {}
    for r in data_rows:
        key = _norm(ws.cell(r, sold_to_col).value)
        observed_sold_to[key] = observed_sold_to.get(key, 0) + 1

    unmapped = {k: v for k, v in observed_sold_to.items() if k and k not in distributor_map_norm}
    if unmapped:
        # Only log top few to avoid noise; still flag it.
        sample = list(unmapped.items())[:8]
        logger.warning(
            "⚠️Found sold-to values in data that are NOT in distributor_map. "
            f"These rows will NOT be written to any distribution sheet. sample={sample}"
        )

    # Create each distributor sheet
    logger.info("Creating distribution tabs...")
    created = 0
    for idx, (sold_to_norm, sheet_title) in enumerate(distributor_map_norm.items(), start=1):
        logger.info(f"[{idx}/{len(distributor_map_norm)}] processing distributor '{sheet_title}'...")
        matching_rows = []
        for r in data_rows:
            v = ws.cell(r, sold_to_col).value
            if _norm(v) == sold_to_norm:
                matching_rows.append(r)

        if not matching_rows:
            logger.info(f"Distributor '{sheet_title}' has 0 matching visible rows; skipping sheet creation.")
            continue

        logger.info(f"Distributor '{sheet_title}' matching rows={len(matching_rows)}.")

        # If rerun, remove existing sheet with same title to keep idempotent behavior
        if sheet_title in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_title}' already exists; deleting and recreating it.")
            del wb[sheet_title]

        dest_name = _safe_sheet_name(wb, sheet_title)
        if dest_name != sheet_title:
            logger.warning(f"⚠️Adjusted sheet name for uniqueness/length: '{sheet_title}' -> '{dest_name}'.")

        dws = wb.create_sheet(dest_name)

        # Column widths
        for c in range(1, ws.max_column + 1):
            w = src_col_widths.get(c)
            if w is not None:
                dws.column_dimensions[get_column_letter(c)].width = w

        # Copy header row (values + styles)
        for c in range(1, ws.max_column + 1):
            _copy_cell(ws.cell(header_row, c), dws.cell(1, c))

        dws.freeze_panes = "A2"

        # Group by contractor (Name of ship-to party), sorted A→Z
        groups: Dict[str, List[int]] = {}
        for r in matching_rows:
            name_val = ws.cell(r, ship_name_col).value
            contractor = "" if name_val is None else str(name_val).strip()
            groups.setdefault(contractor, []).append(r)

        if "" in groups:
            logger.warning(
                f"⚠️Distributor '{sheet_title}' has {len(groups[''])} row(s) with blank contractor name. "
                "These will be grouped under an empty header (still included)."
            )

        contractors_sorted = sorted(groups.keys(), key=lambda x: x.casefold())

        # Style for contractor merged header row
        yellow_fill = PatternFill(patternType="solid", fgColor="FFFF00")
        center = Alignment(horizontal="center", vertical="center")
        bold = Font(bold=True)

        out_r = 2  # start after header
        start_col = min(ship_name_col, name2_col)
        end_col = max(ship_name_col, name2_col)

        # Progress inside distributor
        for g_idx, contractor in enumerate(contractors_sorted, start=1):
            if g_idx == 1 or g_idx % 25 == 0 or g_idx == len(contractors_sorted):
                logger.info(
                    f"Distributor '{sheet_title}' writing contractor {g_idx}/{len(contractors_sorted)}: "
                    f"'{contractor[:60] + ('…' if len(contractor) > 60 else '')}' "
                    f"(rows={len(groups[contractor])})."
                )

            # Two blank lines ABOVE each contractor (even before first group)
            out_r += 1  # upper blank row
            out_r += 1  # lower blank row = merged/highlighted header

            # Merge across Name of ship-to party .. Name 2
            try:
                dws.merge_cells(
                    start_row=out_r,
                    start_column=start_col,
                    end_row=out_r,
                    end_column=end_col,
                )
            except Exception as e:
                logger.warning(
                    f"⚠️Merge failed for distributor '{sheet_title}', contractor '{contractor}' "
                    f"at row={out_r}, cols={start_col}..{end_col}: {e}"
                )

            mcell = dws.cell(out_r, start_col)
            mcell.value = contractor
            mcell.fill = yellow_fill
            mcell.alignment = center
            mcell.font = bold

            # Apply same style to the full merged region cells
            for c in range(start_col, end_col + 1):
                cell = dws.cell(out_r, c)
                cell.fill = yellow_fill
                cell.alignment = center
                cell.font = bold

            # Write contractor rows
            for src_r in groups[contractor]:
                out_r += 1
                src_cells = row_snaps[src_r]
                for c, src_cell in enumerate(src_cells, start=1):
                    _copy_cell(src_cell, dws.cell(out_r, c))

        created += 1
        logger.info(f"Finished sheet '{dest_name}'. rows_written={len(matching_rows)}; final_row={out_r}.")

    logger.info(f"Completed. created_sheets={created}; final_workbook_sheets={wb.sheetnames}.")

    if save:
        out_path = os.path.join(OUTPUT_ROOT, save_name)
        wb.save(out_path)

    return wb


def step_05_create_orders_on_hold_tabs(
        wb: Workbook,
        *,
        source_sheet_name: str = "Sheet1",
        header_scan_rows: int = 20,
        save: bool = False,
        save_name: str = "step5_orders_on_hold.xlsx",
) -> Workbook:
    """
    STEP 5 (Orders on Hold)
    Input: Step 4 output workbook
    Base sheet: Sheet1

    Creates 2 tabs:
      - orders on hold - abc  (East+West separated by red horizontal line)
      - orders on hold - qxo  (East+West separated by red horizontal line)

    Rules:
      - Filter by "Name of sold-to party" using the fixed mappings:
          ABC East = "ABC SUPPLY #402"
          ABC West = "ABC SUPPLY CO INC #015"
          QXO East = "BEACON BUILDING PRODUCTS 285"
          QXO West = "BEACON BUILDING PRODUCTS 228"
      - Keep only rows where "Delivery block description" is NOT blank
      - Sort within each region (East/West) by "Name of ship-to party"
      - Insert blank white row between each job (ship-to party)
      - Insert a red horizontal separator row between East and West
      - Highlight "Delivery block description" column yellow in output tabs
      - If sheet exists already, delete & recreate
    """
    if source_sheet_name not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet_name}' not found in workbook. sheets={wb.sheetnames}")

    ws = wb[source_sheet_name]

    required = ["Name of sold-to party", "Name of ship-to party", "Delivery block description"]
    header_row, col_map = _find_header_row_best_match(ws, required, header_scan_rows=header_scan_rows)

    sold_to_col = col_map["Name of sold-to party"]
    ship_to_col = col_map["Name of ship-to party"]
    del_block_col = col_map["Delivery block description"]

    # Snapshot headers (current sheet order)
    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        headers.append("" if ws.cell(header_row, c).value is None else str(ws.cell(header_row, c).value))

    # Column widths from source
    src_col_widths: Dict[int, Optional[float]] = {}
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        src_col_widths[c] = ws.column_dimensions[letter].width

    # Snapshot row cells for speed (all columns)
    data_rows = list(range(header_row + 1, ws.max_row + 1))
    row_snaps: Dict[int, List] = {}
    for i, r in enumerate(data_rows, start=1):
        row_snaps[r] = [ws.cell(r, c) for c in range(1, ws.max_column + 1)]
        if i % 5000 == 0 or i == len(data_rows):
            logger.info(f"snapshotted {i}/{len(data_rows)} rows.")

    yellow_fill = PatternFill(patternType="solid", fgColor="FFFF00")

    red_fill = PatternFill(patternType="solid", fgColor="FF0000")

    def _is_blank(v) -> bool:
        return _norm(v) == ""

    def _red_separator_row(dws: Worksheet, r: int):
        # draw a solid red bar across the whole row (all columns) using fill
        for c in range(1, ws.max_column + 1):
            cell = dws.cell(r, c)
            cell.value = None
            cell.fill = red_fill

    def _write_orders_tab(sheet_title: str, east_sold_to: str, west_sold_to: str) -> None:
        # delete & recreate if exists
        if sheet_title in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_title}' already exists; deleting and recreating it.")
            del wb[sheet_title]

        dws = wb.create_sheet(sheet_title)

        # widths
        for c in range(1, ws.max_column + 1):
            w = src_col_widths.get(c)
            if w is not None:
                dws.column_dimensions[get_column_letter(c)].width = w

        # header row copy (value + style)
        for c in range(1, ws.max_column + 1):
            _copy_cell(ws.cell(header_row, c), dws.cell(1, c))

        dws.freeze_panes = "A2"

        east_norm = _norm(east_sold_to)
        west_norm = _norm(west_sold_to)

        east_rows: List[int] = []
        west_rows: List[int] = []

        # filter rows by sold-to AND delivery block desc non-blank
        for r in data_rows:
            sold_norm = _norm(ws.cell(r, sold_to_col).value)
            if sold_norm not in (east_norm, west_norm):
                continue

            if _is_blank(ws.cell(r, del_block_col).value):
                continue

            if sold_norm == east_norm:
                east_rows.append(r)
            else:
                west_rows.append(r)

        logger.info(f"'{sheet_title}' rows after filter: east={len(east_rows)}, west={len(west_rows)}")

        def ship_key(rr: int) -> str:
            v = ws.cell(rr, ship_to_col).value
            return ("" if v is None else str(v).strip()).casefold()

        east_rows.sort(key=ship_key)
        west_rows.sort(key=ship_key)

        out_r = 2  # start after header

        def _write_group(rows_list: List[int]):
            nonlocal out_r
            prev_ship = None

            for rr in rows_list:
                ship = "" if ws.cell(rr, ship_to_col).value is None else str(ws.cell(rr, ship_to_col).value).strip()

                # blank row between each job (ship-to party)
                if prev_ship is not None and ship.casefold() != prev_ship.casefold():
                    out_r += 1  # blank white row

                out_r += 1
                for c, src_cell in enumerate(row_snaps[rr], start=1):
                    _copy_cell(src_cell, dws.cell(out_r, c))

                # highlight delivery block description column yellow
                dws.cell(out_r, del_block_col).fill = yellow_fill

                prev_ship = ship

        # Write East then West, separated by red line row (only if both exist)
        if east_rows:
            _write_group(east_rows)

        if east_rows and west_rows:
            out_r += 1  # blank row ABOVE red bar
            out_r += 1
            _red_separator_row(dws, out_r)  # red bar row
            out_r += 1  # blank row BELOW red bar

        if west_rows:
            _write_group(west_rows)

    # Create the two required tabs
    _write_orders_tab(
        "Orders on Hold - ABC (East & West)",
        east_sold_to="ABC SUPPLY #402",
        west_sold_to="ABC SUPPLY CO INC #015",
    )
    _write_orders_tab(
        "Orders on Hold - QXO (East & West)",
        east_sold_to="BEACON BUILDING PRODUCTS 285",
        west_sold_to="BEACON BUILDING PRODUCTS 228",
    )

    if save:
        out_path = os.path.join(OUTPUT_ROOT, save_name)
        wb.save(out_path)

    return wb
